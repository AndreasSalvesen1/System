# pylint: disable=C

import csv
import datetime
import json
import os
import pprint
import shutil
import time
from datetime import datetime, timedelta
from time import perf_counter

import pandas as pd
import requests
from openpyxl import load_workbook
    
# --------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# Programfunksjoner


def regnskapforing():

    # Sjekk om det er ukategoriserte transaksjoner i regnearket Transaksjoner
    # Hvis ja: Fortsett
    # Hvis nei: Gi beskjed om dette og returner

    # Last inn alle kategoriserte transaksjoner i regnearket Transaksjoner til et dictionary
    # Last inn alle ukategoriserte transaksjoner i regnearket Transaksjoner til et dictionary

    # Spør bruker om de vil starte regnskapsføringen

    # Skriv ut alle ukategoriserte transaksjoner som er lagt inn i excel-arket Transaksjoner

    # Etter hver femte transaksjon, spør bruker om de vil fortsette
    # Hvis ja: Fortsett
    # Hvis nei: Etter siste transaksjon, spør bruker om de vil lagre endringene

    # Display 1 ukategorisert transaksjon, og spør bruker om de vil skrive inn en predefinert transaksjon for denne
    # Hvis ja: Skriv inn predefinert transaksjon
    # Dersom input bruker holder på å skrive inn er lik en beskrivelse av en kategorisert transaksjon i regnearket, display denne transaksjonen i en meny over der bruker skriver
    # Display 5 transaksjoner i menyen som maks
    # Ranger dem etter hvor like de er input bruker har skrevet inn
    # Dersom bruker trykker tab skal forslaget som er rangert øverst skrives inn som predefinert transaksjon

    # Hvis nei: Be om at bruker kategoriserer transaksjonen
    # Display alle kategoriene som er lagt inn i regnearket, med tallverdier foran
    # Dersom bruker skriver inn et tall, skal transaksjonen kategoriseres til kategorien som tilsvarer tallet

    # Dersom det ikke er flere transaksjoner igjen, spør bruker om de vil lagre endringene
    # Hvis ja: Lagre endringene og returner
    # Hvis nei: Returner

    return


def hent_transaksjonerSB1():    

    def forsiktig_med_429():

        print(
            "------------------------------------------------------------------------"
        )
        print(
            "Sjekker i Excel om det er gått mer enn",
            pauseperiode_i_timer,
            "timer siden forrige kjøring ...",
        )
        print(
            "------------------------------------------------------------------------"
        )
        NL(1)

        def hent_endringstid_fra_excel():
            change_time = arbeidsark["D5"].value
            if isinstance(change_time, str):
                change_time = datetime.strptime(change_time, "%d.%m.%Y %H:%M:%S")
            return change_time

        endringstid = hent_endringstid_fra_excel()
        if endringstid is not None:
            tid_siden_endring = datetime.now() - endringstid
            ventetid = timedelta(hours=pauseperiode_i_timer)

            if tid_siden_endring < ventetid:
                gjenvaerende_tid = ventetid - tid_siden_endring
                gjenvaerende_tid = gjenvaerende_tid - timedelta(
                    microseconds=gjenvaerende_tid.microseconds
                )
                print(
                    "Det er",
                    gjenvaerende_tid.seconds // 3600,
                    "timer og",
                    (gjenvaerende_tid.seconds % 3600) // 60,
                    "minutter igjen til neste kjøring",
                )
                NL(1)
                if gjenvaerende_tid > timedelta(hours=3):
                    print("Det kan kanskje være lurt å jobbe med C++ imens ...")
                    NL(1)
                print(
                    "Lukker Excel-arbeidsbok slik at bruker også kan redigere i mellomtiden uten at programmet overskriver endringer ..."
                )
                NL(1)
                if gjenvaerende_tid > timedelta(minutes=1):
                    print(
                        "Venter til",
                        (datetime.now() + gjenvaerende_tid).strftime("%H:%M"),
                        "...",
                    )
                    NL(1)
                    # print("Antall sekunder:", gjenvaerende_tid.total_seconds()); NL(1)
                lukk_arbeidsbok_global()
                global arbeidsbok
                time.sleep(gjenvaerende_tid.total_seconds())
                # print("Neste kjøring starter klokken", datetime.now().strftime("%H:%M:%S %Y-%m-%d"), "..."); NL(3)
                print(
                    "Neste kjøring starter klokken",
                    datetime.now().strftime("%H:%M:%S %Y-%m-%d"),
                    "...",
                )
                NL(3)
                arbeidsbok = last_arbeidsbok_global()
            if tid_siden_endring > ventetid:
                print(
                    "Neste kjøring starter klokken",
                    datetime.now().strftime("%H:%M:%S %Y-%m-%d"),
                    "...",
                )
                NL(3)

    def hent_transaksjoner():

        start_tid = perf_counter()
        print(
            "----------------------------------------------------------------------------------------------------------------------------------"
        )
        print(
            "--------------------------------------- Starter henting av transaksjoner fra SpareBank1 ... --------------------------------------"
        )
        print(
            "----------------------------------------------------------------------------------------------------------------------------------"
        )
        NL(3)
        time.sleep(3)

        def hent_kontoer():
            start_tid2 = perf_counter()
            print("-----------------------")
            print("Henter alle kontoer ...")
            print("-----------------------")
            NL(1)
            paramatere = "includeNokAccounts=true&includeCurrencyAccounts=true&includeBsuAccounts=true&includeHiddenAccounts=true&includeCreditCardAccounts=true&includeAskAccounts=true&includePensionAccounts=true"
            request_URL = (
                "https://api.sparebank1.no/personal/banking/accounts?" + paramatere
            )
            reqheadere = {
                "Authorization": "Bearer " + open(ACCESS_TOKEN_FILE).read().strip(),
                "Accept": "application/vnd.sparebank1.v1+json;charset=utf-8",
            }
            SB1_output = requests.get(request_URL, headers=reqheadere)
            data = SB1_output.json()
            kontonummere_SB1 = [konto["accountNumber"] for konto in data["accounts"]]
            print(
                "Ferdig med henting av kontoer etter:",
                round(perf_counter() - start_tid2, 3),
                "sekunder",
            )
            NL(1)
            # print(json.dumps(data, indent=4)); NL(1)
            print("Sparebank1 sier at kontonummere er:")
            print(kontonummere_SB1)
            NL(3)

        def hent_access_keys():
            start_tid2 = perf_counter()
            print("-----------------------")
            print("Henter account keys ...")
            print("-----------------------")
            NL(1)
            SB1_base_url = "https://api.sparebank1.no/personal/banking/accounts"
            # kontonummere = hent_kontoer()
            parametere = "&".join(f"accountNumber={Konto}" for Konto in kontonummere)

            reqheadere = {
                "accept": "application/vnd.sparebank1.v1+json;charset=utf-8",
                "Authorization": f"Bearer {open(ACCESS_TOKEN_FILE).read().strip()}",
            }

            response = requests.get(
                f"{SB1_base_url}/keys?{parametere}", headers=reqheadere
            )
            data = response.json()
            print(
                "Ferdig med henting av account keys etter:",
                round(perf_counter() - start_tid2, 3),
                "sekunder",
            )
            NL(1)
            print("Account keys er:")
            NL(1)
            for account_number, account_key in data["accountKeysMap"].items():
                print("Kontonummer:", account_number, "------>", account_key)
            NL(3)

            with open(AccountKeyMap_Fil, "w") as file:
                json.dump(data, file)

        def refresh_access_token():
            print("----------------------------------------------")
            print("Oppdaterer access token, med refresh token ...")
            print("----------------------------------------------")
            NL(1)
            start_tid2 = perf_counter()

            token_adresse = "https://api-auth.sparebank1.no/oauth/token"
            client_id = CLIENT_ID
            client_secret = CLIENT_SECRET

            with open(REFRESH_TOKEN_FILE, "r") as file:
                refresh_token = file.read().strip()
                print("Refresh token før request er:", refresh_token)
                NL(1)

            data = {
                "client_id": client_id,
                "client_secret": client_secret,
                "refresh_token": refresh_token,
                "grant_type": "refresh_token",
            }

            print("Request til SB1:")
            print(f"POST {token_adresse}")
            NL(1)
            print("Parametere er:")
            pprint.pprint(data)
            NL(1)

            SB1_output = requests.post(token_adresse, data=data)
            Output_data = SB1_output.json()
            print("Statuskode for request er:", SB1_output.status_code)
            print(
                "Ferdig med oppdatering av access token etter:",
                round(perf_counter() - start_tid2, 3),
                "sekunder",
            )
            NL(1)
            # print("Respons fra SpareBank1 er:", Output_data); print()
            # print("Access token etter request er:", response_data["access_token"]); print()
            try: 
                print("Refresh token etter request er:", Output_data["refresh_token"])
            except KeyError:
                print("Output fra SB1 ble:", str(Output_data))
            NL(1)

            with open(ACCESS_TOKEN_FILE, "w") as file:
                file.write(Output_data["access_token"])

            with open(REFRESH_TOKEN_FILE, "w") as file:
                file.write(Output_data["refresh_token"])

        def hent_transaksjoner(Kontonummer):
            start_tid2 = perf_counter()
            # til = '2023-11-11'
            til = oppdater_til_dato
            fra = oppdater_fra_dato
            print(
                "----------------------------------------------------------------------------------------------------------------"
            )
            print(
                "Starter henting av transaksjoner fra SpareBank1 for kontonummer:",
                Kontonummer,
                f"fra: {fra} til: {til} ...",
            )
            print(
                "----------------------------------------------------------------------------------------------------------------"
            )
            NL(1)

            Eksisterende_CSV_fil = os.path.join(H_katalog, f"{Kontonummer}.csv")
            Skriv_til_CSV_fil = os.path.join(path_til_CSV, f"{Kontonummer}.csv")

            with open(AccountKeyMap_Fil, "r") as file:
                try:
                    a_key_map = json.load(file)
                except json.JSONDecodeError:
                    a_key_map = {}

            a_key = a_key_map["accountKeysMap"].get(Kontonummer)
            if a_key is None:
                print(f"Kunne ikke finne account key for kontonummeret {Kontonummer}")
            else:
                print(f"Account key for kontonummeret {Kontonummer}: {a_key}")

            url = f"https://api.sparebank1.no/personal/banking/transactions/export?accountKey={a_key}&fromDate={fra}&toDate={til}"
            headers = {
                "Accept": "application/csv;charset=utf-8",
                "Authorization": f"Bearer {open(ACCESS_TOKEN_FILE).read().strip()}",
            }
            # print("Headers er:", headers)
            # print("URL er:", url)
            SB1_output = requests.get(url, headers=headers)

            if os.path.exists(Eksisterende_CSV_fil):
                Eksisterende_CSV_data = []
                with open(Eksisterende_CSV_fil, "r") as f:
                    Les = csv.reader(f)
                    Eksisterende_CSV_data = list(Les)

                ny_data = SB1_output.text.splitlines()
                Eksisterende_CSV_Dato = [row[0] for row in Eksisterende_CSV_data]
                nye_datoer = [row.split(",")[0] for row in ny_data]
                overlappende_datoer = set(Eksisterende_CSV_Dato) & set(nye_datoer)
                filtrert_data = [
                    row
                    for row in Eksisterende_CSV_data
                    if row[0] not in overlappende_datoer
                ]

                with open(Eksisterende_CSV_fil, "w", newline="", encoding="utf-8") as f:
                    writer = csv.writer(f)
                    writer.writerows(filtrert_data)
                    writer.writerows(csv.reader(ny_data))
                print(
                    f"Oppdaterte {Eksisterende_CSV_fil} med {len(ny_data)} nye transaksjoner"
                )

            else:
                with open(Eksisterende_CSV_fil, "w", encoding="utf-8") as f:
                    f.write(SB1_output.text)
                NL(1)
                print(
                    f"Lagret {Eksisterende_CSV_fil} med {len(SB1_output.text.splitlines())} transaksjoner"
                )
                NL(1)

            print(f"Request for transaksjoner i kontonummer: {Kontonummer}:")
            print("-----------------------------------------------------")
            print(url, headers)
            NL(1)
            print("Statuskode for request er:", SB1_output.status_code)
            print("Ny refresh_token er:", open(REFRESH_TOKEN_FILE).read().strip())
            print("-----------------------------------------------------")
            NL(1)
            print("Respons fra SpareBank1 er:")
            NL(1)
            # print("-------------------------------------------------------------------------------------------------------")
            print(SB1_output.text)
            NL(1)

            if os.path.exists(Skriv_til_CSV_fil):
                os.remove(Skriv_til_CSV_fil)
            shutil.move(Eksisterende_CSV_fil, path_til_CSV)
            print(
                "-----------------------------------------------------------------------------------"
            )
            print(
                "Ferdig med henting av transaksjoner i kontonummer",
                Kontonummer,
                "etter:",
                round(perf_counter() - start_tid2, 3),
                "sekunder",
            )
            print(
                "-----------------------------------------------------------------------------------"
            )
            NL(3)

        def HW_test():
            print("--------------------------------")
            print("Starter SB1 Hello-World test ...")
            print("--------------------------------")
            NL(1)
            reqheadere = {
                "Authorization": "Bearer " + open(ACCESS_TOKEN_FILE).read().strip(),
                "Accept": "application/vnd.sparebank1.v1+json;charset=utf-8",
            }
            SB1_output = requests.get(
                "https://api.sparebank1.no/common/helloworld", headers=reqheadere
            )
            data = SB1_output.json()
            print("Respons fra SpareBank1 er:")
            print(json.dumps(data, indent=4))
            melding = data["message"]
            if melding == "Hello, World my friend!":
                print(
                    "SB1 Hello-World test OK, access_token OK, fortsetter til neste steg ..."
                )
                NL(3)
            if melding != "Hello, World my friend!":
                print(
                    "SB1 Hello-World test feilet, access_token feil, avslutter program ..."
                )
                NL(3)
                exit()

        def hent_alle_transaksjoner():
            Kontonummere = kontonummere
            # Kontonummere = ['18131796605']
            for Kontonummer in Kontonummere:
                hent_transaksjoner(Kontonummer)
                time.sleep(5)

        def start_funksjon():
            functions_to_run = [
                refresh_access_token,
                HW_test,
                hent_kontoer,
                hent_access_keys,
                hent_alle_transaksjoner
            ]

            for func in functions_to_run:
                try:
                    func()
                except Exception as e:
                    print(f"Funksjonen {func.__name__} feilet", ); NL(2)
                    return
                time.sleep(3)
            # print('Alle kontoer i', hovedkatalog,'behandlet i Sparebank1 API, fortsetter nå til videre prosessering ...')

        try:
            start_funksjon()
        except Exception as e:
            NL(1)
            # print(e); NL(1)
            print(
                "Fortsetter programmet likevel og prøver å hente transaksjoner igjen om 8 timer ..."
            )
            NL(4)

        print(
            "----------------------------------------------------------------------------------------------------------------------------------"
        )
        print(
            "-------------------------- Ferdig med henting av transaksjoner fra SpareBank1 etter:",
            round(perf_counter() - start_tid, 3),
            "sekunder -----------------------------",
        )
        print(
            "----------------------------------------------------------------------------------------------------------------------------------"
        )
        NL(3)

    forsiktig_med_429()
    time.sleep(4)

    hent_transaksjoner()
    time.sleep(4)


def last_arbeidsbok_global():
    NL(2)
    path_til_Excel = os.path.normpath(os.path.join(hovedkatalog, Excel_filnavn))
    # print("hoventkatalog er:", hovedkatalog)
    # print("path_til_Excel er:", path_til_Excel)
    print("------------------------------------------")
    print("Starter innlasting av Excel-arbeidsbok ...")
    print("------------------------------------------")
    start_tid = perf_counter()
    while True:
        try:
            arbeidsbok = load_workbook(path_til_Excel)
            break
        except PermissionError as e:
            venteperiode = 60 * 60 * 1.5
            tid = datetime.now() + timedelta(seconds=venteperiode)
            NL(1)
            print("Excel-fil var nå åpen, dette gir følgende feilmelding:")
            print(e)
            NL(1)
            gjenvaerende_tid = venteperiode
            timer = gjenvaerende_tid // 3600
            minutter = round(gjenvaerende_tid % 3600 // 60, 0)
            print(
                "Venteperiode er satt til",
                round(timer, None),
                "time og",
                round(minutter, None),
                "minutter før neste forsøk ...",
            )
            NL(1)
            print("Venter til:", tid.strftime("%H:%M"), "for å prøve på nytt ...")
            time.sleep(venteperiode)
    NL(1)
    print("----------------------------------------------------------------")
    print(
        "Ferdig med innlasting av Excel-arbeidsbok, etter:",
        round(perf_counter() - start_tid, 3),
        "sekunder",
    )
    print("----------------------------------------------------------------")
    NL(1)
    return arbeidsbok


def fjern_all_data_BtilG():
    print("--------------------------------------------------")
    print("Fjerner data i kolonne B-G fra rad 9 og nedover...")
    print("--------------------------------------------------")
    NL(1)
    start_tid = perf_counter()
    start_rad = 9
    slutt_rad = arbeidsark.max_row
    start_kolonne = 2
    slutt_kolonne = 7

    for row in range(start_rad, slutt_rad + 1):
        for col in range(start_kolonne, slutt_kolonne + 1):
            arbeidsark.cell(row=row, column=col).value = None

    print(
        "Ferdig med fjerning av data fra B-F etter:",
        round(perf_counter() - start_tid, 3),
        "sekunder",
    )
    NL(3)


def fjern_overlappende_data_BtilG():

    print(
        "-----------------------------------------------------------------------------------------------------"
    )
    print(
        "Fjerner data i kolonne B-G fra rad 9 og nedover dersom dato i kolonne B er høyere enn:",
        oppdater_fra_dato,
        "...",
    )
    print(
        "-----------------------------------------------------------------------------------------------------"
    )
    NL(1)
    start_tid = perf_counter()
    start_rad = 9
    slutt_rad = arbeidsark.max_row
    start_kolonne = 2
    slutt_kolonne = 7
    global antall_tomme_rader

    csv_datoer = []
    for csv_filnavn in CSV_filnavn:
        csv_filbane = os.path.join(path_til_CSV, csv_filnavn)
        print("CSV-filbane er:", csv_filbane)
        with open(csv_filbane, "r") as csv_file:
            csv_reader = csv.reader(csv_file)
            next(csv_reader)  # Skip header row
            for row in csv_reader:
                csv_date = row[0].split(";")[0].strip('"')
                csv_datoer.append(csv_date)
    NL(1)
    print("Alle datoene i CSV-filen er:")
    NL(1)
    pprint.pprint(set(csv_datoer))
    NL(1)

    print(
        "Fjerner data fra Excel hvis B-kolonne inneholder en dato som også er en av disse CSV-datoene ..."
    )
    NL(1)
    for row in range(start_rad, slutt_rad + 1):
        excel_dato = arbeidsark.cell(row=row, column=start_kolonne).value
        if excel_dato in csv_datoer:
            antall_tomme_rader += 1
            for col in range(start_kolonne, slutt_kolonne + 1):
                arbeidsark.cell(row=row, column=col).value = None

    if antall_tomme_rader != 0:
        print("Antall oppdaterte rader er:", antall_tomme_rader)

    NL(1)
    print(
        "Ferdig med fjerning av data fra B-F etter:",
        round(perf_counter() - start_tid, 3),
        "sekunder",
    )
    NL(3)


def hent_neste_tomme_rad(kolonnestart, kolonneslutt, startradnummer):
    radnummer = startradnummer
    # print("Starter på rad:", radnummer)
    none_counter = 0
    while True:
        if all(
            arbeidsark.cell(row=radnummer, column=kolonnenummer).value is None
            for kolonnenummer in range(kolonnestart, kolonneslutt)
        ):
            none_counter += 1
            # print("Neste tomme rad er:", radnummer); NL(1)
            # return radnummer
        else:
            none_counter = 0

        if none_counter == 100:
            radnummer -= 99
            # print("Neste tomme rad er:", radnummer); NL(1)
            return radnummer

        radnummer += 1


def hent_neste_ikke_tomme_rad(kolonnestart, kolonneslutt, startradnummer):
    radnummer = startradnummer
    # print("Starter på rad:", radnummer)
    not_none_counter = 0
    while True:
        if all(
            arbeidsark.cell(row=radnummer, column=kolonnenummer).value is not None
            for kolonnenummer in range(kolonnestart, kolonneslutt)
        ):
            not_none_counter += 1
        else:
            not_none_counter = 0

        if not_none_counter == 100:
            radnummer -= 99
            # print("Neste tomme rad er:", radnummer); NL(1)
            return radnummer

        radnummer += 1


def last_kategoriseringsregler():
    print("Starter innlasting av kategoriseringsregler fra markert seksjon ...")
    NL(1)
    start_tid = perf_counter()

    kategoriark = arbeidsark
    kategoriseringsregler = {}

    for kolonne in range(9, kategoriark.max_column + 1):
        kategori = kategoriark.cell(row=2, column=kolonne).value
        if kategori is not None:
            print(f"Kolonne {kolonne}, Kategori: {kategori} ferdig ...")
            time.sleep(0.5)

        if kategori is None:
            print(
                f"Stopper ved kolonne: {kolonne} fordi kategori er definert som None."
            )
            break

        # siste_rad = 3
        # for rad in range(kategoriark.max_row, 2, -1):
        #     if kategoriark.cell(row=rad, column=kolonne).value is not None:
        #         siste_rad = rad
        #         break

        # # print(f"Siste rad for data i kolonne {kolonne} er {siste_rad}")

        for rad in range(3, hent_neste_tomme_rad(kolonne, kolonne + 1, 3)):
            indikator = kategoriark.cell(row=rad, column=kolonne).value
            if indikator is not None and indikator != "":
                kategoriseringsregler[indikator] = kategori

    NL(1)
    print(
        f"Ferdig med innlasting av kategoriseringsregler etter: {round(perf_counter() - start_tid, 3)} sekunder"
    )
    NL(1)
    # print("Kategoriseringsregler er:")
    # print(kategoriseringsregler); NL(1)
    return kategoriseringsregler


def kategoriser_transaksjoner(beskrivelse, kategoriseringsregler):
    for indikator, kategori in kategoriseringsregler.items():
        if indikator in str(beskrivelse):
            return kategori
    return "Ukjent"


def hent_kategoriser_og_skriv_CSV():

    print("--------------------------------------------------------")
    print("Starter prosessering av alle CSV-filer til dataframe ...")
    print("--------------------------------------------------------")
    NL(1)
    time.sleep(5)

    start_tid2 = perf_counter()
    global neste_tomme_rad
    neste_tomme_rad = hent_neste_tomme_rad(2, 5, 9)
    kategoriseringsregler = last_kategoriseringsregler()
    for CSV_fil in CSV_filnavn:
        start_tid = perf_counter()
        print("Starter prosessering av CSV-fil", CSV_fil, "...")

        # path_til_CSV = os.path.normpath(os.path.join(path_til_CSV, CSV_fil))
        if not os.path.exists(path_til_CSV):
            print(f"CSV-fil ikke funnet: {path_til_CSV}")
            continue

        filpath = os.path.join(path_til_CSV, CSV_fil)
        df = pd.read_csv(filpath, sep=";")
        df["Inn"] = pd.to_numeric(
            df["Inn"].astype(str).str.replace(",", ".").replace("-", ""),
            errors="coerce",
        ).fillna(0)
        df["Ut"] = pd.to_numeric(
            df["Ut"].astype(str).str.replace(",", ".").replace("-", ""), errors="coerce"
        ).fillna(0)
        df["Mengde"] = df["Inn"] + df["Ut"]
        df = df.drop(columns=["Inn", "Ut"])
        # NL(5); print("Linje 467"); exit()
        df = df[
            ~df["Beskrivelse"]
            .astype(str)
            .str.contains("|".join(eksluder_disse_beskrivelser))
            .fillna(False)
        ]

        print("Transaksjoner med noen av disse ordene i beskrivelsene ekskluderes:")
        print(eksluder_disse_beskrivelser)

        print("Starter kategorisering av transaksjoner ...")
        start_tid1 = perf_counter()
        df["Kategori"] = df["Beskrivelse"].apply(
            lambda x: kategoriser_transaksjoner(x, kategoriseringsregler)
        )
        df = df[["Dato", "Beskrivelse", "Kategori", "Mengde"]]
        print(
            "Ferdig med kategorisering etter:",
            round(perf_counter() - start_tid1, 3),
            "sekunder",
        )

        antall_transaksjoner = 0
        print(
            "Starter skriving av ny data tilbake til Excel fra",
            CSV_fil,
            "ved rad:",
            neste_tomme_rad,
            "...",
        )
        start_write = perf_counter()
        for row_data in df.values.tolist():
            arbeidsark.cell(row=neste_tomme_rad, column=2, value=row_data[0])
            arbeidsark.cell(row=neste_tomme_rad, column=3, value=row_data[1])
            arbeidsark.cell(row=neste_tomme_rad, column=4, value=row_data[2])
            arbeidsark.cell(row=neste_tomme_rad, column=5, value=row_data[3])
            neste_tomme_rad += 1
            antall_transaksjoner += 1

        print(
            "Ferdig med skriving av ny CSV tilbake til Excel etter:",
            round(perf_counter() - start_write, 3),
            "sekunder",
        )
        print(
            "Ferdig med prosessering av CSV-fil",
            CSV_fil,
            "etter:",
            round(perf_counter() - start_tid, 3),
            "sekunder, med",
            antall_transaksjoner,
            "transaksjoner",
        )
        NL(1)

    # print("----------------------------------------------------------------")
    print(
        "Ferdig med prosessering av alle CSV-filer etter:",
        round(perf_counter() - start_tid2, 3),
        "sekunder",
    )
    # print("----------------------------------------------------------------")
    NL(3)


def skriv_predefinerte_transaksjoner():

    def skriv_nye_beskrivelser():
        time.sleep(3)
        print("-----------------------------------------------------------")
        print("Starter oppdatering av beskrivelser fra markert seksjon ...")
        print("-----------------------------------------------------------")
        NL(1)
        start_tid = perf_counter()

        til1 = hent_neste_tomme_rad(
            23 + antall_nye_kolonner, 26 + antall_nye_kolonner, 3
        )
        print("Siste radnummer i sammenligningsdel er:", til1 - 1)
        NL(1)

        til2 = hent_neste_tomme_rad(2, 5, 9)
        # til2 = 20
        print("Siste radnummer i transaksjonsdel er:", til2 - 1)
        NL(1)

        antall_endringer = 0

        for rad in range(3, til1):
            # print("Rad i sammenligningsdel er:", rad)

            mengdeindikator = arbeidsark.cell(
                row=rad, column=(24 + antall_nye_kolonner)
            ).value
            indikativ_beskrivelse = arbeidsark.cell(
                row=rad, column=(25 + antall_nye_kolonner)
            ).value
            ny_beskrivelse = arbeidsark.cell(
                row=rad, column=(26 + antall_nye_kolonner)
            ).value
            # print("Ny beskrivelser er:", ny_beskrivelse)

            for sammenlign_rad in range(9, til2):
                # print("Sammenlign_rad er:", sammenlign_rad)
                sammenlign_med_beskrivelse = arbeidsark.cell(
                    row=sammenlign_rad, column=3
                ).value
                sammenlign_med_mengdeindikator = arbeidsark.cell(
                    row=sammenlign_rad, column=5
                ).value
                # print("Sammenlign med mengdeindikator er:", sammenlign_med_mengdeindikator)
                # print("Sammenlign_med_beskrivelse er:", sammenlign_med_beskrivelse)
                # print("Indikativ_beskrivelse er:", indikativ_beskrivelse)
                # if sammenlign_rad == 9:
                # print("Sammenlign med beskrivelse er:", sammenlign_med_beskrivelse)
                #    print("Sammenlign med mengdeindikator er:", sammenlign_med_mengdeindikator)
                #    print("Indikativ beskrivelse er:", indikativ_beskrivelse)
                #    print("Mengdeindikator er:", mengdeindikator)
                if mengdeindikator is not None:
                    if mengdeindikator == sammenlign_med_mengdeindikator and str(
                        indikativ_beskrivelse
                    ) in str(sammenlign_med_beskrivelse):
                        # print("Kolonne 24 rad", rad ,"er lik kolonne 5 rad", sammenlign_rad, "og kolonne 23 rad", rad, "er lik kolonne 3 rad", sammenlign_rad)
                        # print("Ny beskrivelse er:", ny_beskrivelse)
                        # print("Skriver ny beskrivelse til kolonne 3 rad", sammenlign_rad, "..."); NL(1)
                        arbeidsark.cell(row=sammenlign_rad, column=3).value = (
                            ny_beskrivelse
                        )
                        antall_endringer += 1
                if str(indikativ_beskrivelse) in str(sammenlign_med_beskrivelse):
                    # print("Kolonne 23 rad", rad, "er lik kolonne 3 rad", sammenlign_rad)
                    # print("Ny beskrivelse er:", ny_beskrivelse)
                    # print("Skriver ny beskrivelse til kolonne 3 rad", sammenlign_rad, "..."); NL(1)
                    arbeidsark.cell(row=sammenlign_rad, column=3).value = ny_beskrivelse
                    antall_endringer += 1

        print(
            f"Ferdig med omskriving av",
            antall_endringer,
            "transaksjoner i kolonne C fra rad 9 til rad",
            til2 - 1,
            "etter:",
            round(perf_counter() - start_tid, 3),
            "sekunder",
        )
        NL(1)

        print(
            "Ferdig med rekategorisering av alle transaksjoner etter:",
            round(perf_counter() - start_tid, 3),
            "sekunder",
        )
        NL(3)

    def rekategoriser_alle_transaksjoner():
        print("--------------------------------------------------")
        print("Starter rekategorisering av alle transaksjoner ...")
        print("--------------------------------------------------")
        NL(1)
        start_tid = perf_counter()
        kategoriseringsregler = last_kategoriseringsregler()

        print(
            "Begynner rekategoriserering fra rad 9 til rad",
            hent_neste_tomme_rad(2, 5, 9) - 1,
            "... ",
        )
        NL(1)
        antall_kategoriseringer = 0
        for row in arbeidsark.iter_rows(
            min_row=8, min_col=3, max_col=3, max_row=hent_neste_tomme_rad(2, 5, 9)
        ):
            celle_verdi = row[0].value  # Added .value to get the cell's value
            # print("Cell value:", celle_verdi)  # Debug print
            if isinstance(celle_verdi, str) and celle_verdi.strip():
                d_verdi = celle_verdi.strip()
                kategori = kategoriseringsregler.get(d_verdi)
                # print("Category:", kategori)  # Debug print
                if kategori:
                    arbeidsark.cell(row=row[0].row, column=4).value = kategori
                    # print("Kategoriserte transaksjon i rad", row[0].row, "til kategori:", kategori)
                    antall_kategoriseringer += 1

        print(
            "Ferdig med rekategorisering av",
            antall_kategoriseringer,
            "transaksjoner etter:",
            round(perf_counter() - start_tid, 3),
            "sekunder",
        )
        NL(3)

    skriv_nye_beskrivelser()

    rekategoriser_alle_transaksjoner()


def les_og_sorter():

    print("---------------------------------------")
    print("Starter lesing og sortering av data ...")
    print("---------------------------------------")
    NL(1)

    start_tid = perf_counter()
    max_rad = hent_neste_tomme_rad(2, 5, 9)
    print("Konverterer korrekt Excel-område til Pandas dataframe ...")
    data = pd.DataFrame(
        arbeidsark.iter_rows(
            min_row=9, min_col=2, max_col=7, max_row=max_rad, values_only=True
        )
    )
    # data = pd.DataFrame(arbeidsark.iter_rows(min_row=9, min_col=2, max_col=7, values_only=True))
    data.columns = ["Dato", "Beskrivelse", "Kategori", "Mengde", "Saldo", "Type"]
    try:
        data["Dato"] = pd.to_datetime(data["Dato"], format="%d.%m.%Y")
    except ValueError as e:
        print("Feil ved konvertering av dato:", e)
        return

    if data["Dato"].dtype != "datetime64[ns]":
        print("Feil: 'Dato' kolonnen er ikke i datetime-format.")
        return

    data.sort_values(by=["Dato", "Beskrivelse"], ascending=[False, True], inplace=True)
    data.reset_index(drop=True, inplace=True)
    print(
        "Ferdig med sortering av data etter:",
        round(perf_counter() - start_tid, 3),
        "sekunder",
    )

    print("Konverterer data tilbake til original form ...")
    start_tid = perf_counter()

    data["Dato"] = data["Dato"].dt.strftime("%d.%m.%Y")

    print(
        "Ferdig med konvertering av data tilbake til original form etter:",
        round(perf_counter() - start_tid, 3),
        "sekunder",
    )
    NL(1)

    print("Starter tilbakeskriving av data ...")
    start_tid = perf_counter()

    for row_index, row in data.iterrows():
        for col_index, value in enumerate(row):
            cell = arbeidsark.cell(row=row_index + 9, column=col_index + 2)
            if col_index == 0 and pd.notnull(value):
                cell.number_format = "DD.MM.YYYY"
                cell.value = value
            else:
                cell.value = value

    NL(1)
    print(
        "Ferdig med tilbakeskriving av data etter:",
        round(perf_counter() - start_tid, 3),
        "sekunder",
    )
    NL(3)


def beregn_og_skriv_saldo():
    print("------------------------------------")
    print("Beregner og skriver saldoverdier ...")
    print("------------------------------------")
    NL(1)
    start_tid = perf_counter()
    startsaldo1 = startsaldo
    # antall_tomme_saldo_rader = hent_neste_ikke_tomme_rad(7, 7, 9)
    siste_tomme_rad = hent_neste_tomme_rad(2, 7, 9)

    for row_number in range(siste_tomme_rad - 1, 8, -1):
        cell_value = arbeidsark.cell(row=row_number, column=5).value

        if cell_value is not None and isinstance(cell_value, (int, float)):
            try:
                startsaldo1 += int(cell_value)
            except ValueError:
                pass

        cell = arbeidsark.cell(row=row_number, column=6)
        cell.value = int(startsaldo1)
        # cell.number_format = 'General'

    print("Ferdig med beregning av saldo, skriver tilbake ...")
    NL(1)
    for row_number in range(9, siste_tomme_rad):
        cell_value = arbeidsark.cell(row=row_number, column=6).value

        if cell_value == 0:
            arbeidsark.cell(row=row_number, column=6).value = None

    print(
        "Fjerner bereregnet saldo for tomme rader fra radnummer",
        9,
        "til radnummer",
        siste_tomme_rad - 1,
        "...",
    )
    NL(1)
    for row_number in range(9, siste_tomme_rad):
        cell_value = arbeidsark.cell(row=row_number, column=3).value

        if cell_value is None or cell_value == "":
            # print("Fjerner bereregnet saldo for tom rad nummer:", row_number)
            arbeidsark.cell(row=row_number, column=6).value = None

    print(
        "Ferdig med beregning og tilbake skriving av saldo etter:",
        round(perf_counter() - start_tid, 3),
        "sekunder",
    )
    NL(3)


def skriv_transaksjonstyper():
    start_tid = perf_counter()
    print("-----------------------------------------")
    print("Starter skriving av transaksjonstyper ...")
    print("-----------------------------------------")
    NL(1)
    s_verdier = {
        s_celle[0]
        for s_celle in arbeidsark.iter_rows(
            min_row=3,
            max_row=99,
            min_col=(20 + antall_nye_kolonner),
            max_col=(20 + antall_nye_kolonner),
            values_only=True,
        )
        if s_celle[0] is not None
    }
    t_verdier = {
        t_celle[0]
        for t_celle in arbeidsark.iter_rows(
            min_row=3,
            max_row=99,
            min_col=(21 + antall_nye_kolonner),
            max_col=(21 + antall_nye_kolonner),
            values_only=True,
        )
        if t_celle[0] is not None
    }
    u_verdier = {
        u_celle[0]
        for u_celle in arbeidsark.iter_rows(
            min_row=3,
            max_row=99,
            min_col=(22 + antall_nye_kolonner),
            max_col=(22 + antall_nye_kolonner),
            values_only=True,
        )
        if u_celle[0] is not None
    }

    print("Inntekter er:")
    print(s_verdier)
    NL(1)
    time.sleep(1)
    print("Utgifter er:")
    print(t_verdier)
    NL(1)
    time.sleep(1)
    print("Sparing er:")
    print(u_verdier)
    NL(1)
    time.sleep(1)

    # print("Starter skriving av transaksjonstyper ..."); NL(1)
    for rad in arbeidsark.iter_rows(min_row=9, min_col=4, max_col=7):
        celle_verdi = rad[0].value

        if isinstance(celle_verdi, str) and celle_verdi.strip():
            d_verdi = celle_verdi.strip()

            if d_verdi in t_verdier:
                arbeidsark.cell(row=rad[0].row, column=7).value = "Utgifter"
            elif d_verdi in s_verdier and d_verdi not in u_verdier:
                arbeidsark.cell(row=rad[0].row, column=7).value = "Inntekt"
            elif d_verdi in u_verdier:
                arbeidsark.cell(row=rad[0].row, column=7).value = "Sparing"
            else:
                arbeidsark.cell(row=rad[0].row, column=7).value = ""
        else:
            break
    print(
        "Ferdig med skriving av transaksjonstyper etter:",
        round(perf_counter() - start_tid, 3),
        "sekunder",
    )
    NL(3)


def skriv_endringstid():
    rad = 5
    kolonne = 4
    print("---------------------------")
    print("Registrerer endringstid ...")
    print("---------------------------")
    NL(1)
    now = datetime.now()
    endringsdato = now.strftime("%d.%m.%Y %H:%M:%S")
    arbeidsark.cell(row=rad, column=kolonne).value = None
    arbeidsark.cell(row=rad, column=kolonne).value = endringsdato
    cell_value = arbeidsark["D3"].value
    arbeidsark["D3"].value = cell_value + 1
    print("Endringstid er registrert som:", endringsdato)
    NL(3)
    arbeidsark.cell(row=6, column=4).value = arbeidsark.cell(row=9, column=6).value


def lagre():

    print("-----------------------------------------")
    print("Lagrer endringer, ikke avbryt kjøring ...")
    print("-----------------------------------------")
    NL(1)
    start_tid = perf_counter()
    while True:
        try:
            arbeidsbok.save(path_til_Excel)
            lukk_arbeidsbok_global()
            break
        except PermissionError as e:
            venteperiode = 60 * 60 * 0.25
            tid = datetime.now() + timedelta(seconds=venteperiode)
            NL(1)
            print("Excel-fil var nå åpen, dette gir følgende feilmelding:")
            print(e)
            NL(1)
            gjenvaerende_tid = venteperiode
            timer = round(gjenvaerende_tid // 3600, 0)
            minutter = round(gjenvaerende_tid % 3600 // 60, 0)
            print(
                "Venteperiode er satt til",
                round(timer, None),
                "time og",
                round(minutter, None),
                "minutter før neste forsøk ...",
            )
            NL(1)
            print("Venter til:", tid.strftime("%H:%M"), "for å prøve på nytt ...")
            time.sleep(venteperiode)

    print(
        "Ny data fra CSV-filer i mappen:",
        path_til_CSV,
        "er lagret i Excel-filen:",
        Excel_filnavn,
        "etter:",
        round(perf_counter() - start_tid, 3),
        "sekunder",
    )
    NL(3)
    print(
        "----------------------------------------------------------------------------------------------------------------------------------"
    )
    print(
        "---------------------------------------------- Kode fullført etter",
        round(perf_counter() - total_tid_variabel, 3),
        "sekunder ----------------------------------------------",
    )
    print(
        "----------------------------------------------------------------------------------------------------------------------------------"
    )
    NL(30)


def printStartInfo():

    global CLIENT_ID
    global CLIENT_SECRET
    global oppdater_fra_dato
    global oppdater_til_dato
    global CSV_filnavn
    global path_til_Excel
    global total_tid_variabel
    global neste_tomme_rad
    global antall_tomme_rader

    NL(1)
    # telle_variabel += 1
    print(
        "----------------------------------------------------------------------------------------------------------------------------------"
    )
    print(
        "---------------------------------------------------- Starter kjøring nummer",
        telle_variabel,
        "----------------------------------------------------",
    )
    print(
        "----------------------------------------------------------------------------------------------------------------------------------"
    )
    NL(2)
    total_tid_variabel = perf_counter()
    osenv = os.name
    neste_tomme_rad = None
    antall_tomme_rader = 0

    if osenv == "nt":
        print("Registrert operativsystem er Windows")
    if osenv == "posix":
        print("Registrert operativsystem er Linux")
        NL(1)

    print("H-katalog er:", H_katalog)
    NL(1)

    with open(H_katalog + "client_ID.txt", "r") as file:
        CLIENT_ID = file.read().strip()
        print("CLIENT_ID er:", CLIENT_ID)
        if type(CLIENT_ID) != str or len(CLIENT_ID) == 0:
            print("CLIENT_ID er ikke regisrert som streng, avslutter")
            NL(3)
            exit()

    with open(H_katalog + "client_secret.txt", "r") as file:
        CLIENT_SECRET = file.read().strip()
        print("CLIENT SECRET er:", CLIENT_SECRET)
        if type(CLIENT_SECRET) != str or len(CLIENT_SECRET) == 0:
            print("CLIENT_SECRET er ikke registrert som streng, avslutter")
            NL(3)
            exit()

    if type(CLIENT_SECRET) == str and type(CLIENT_ID) == str:
        NL(1)
        print(
            "CLIENT_ID og CLIENT_SECRET er begge registrert som strenger, fortsetter ..."
        )
        NL(1)

    print("Path til CSV er:", path_til_CSV)
    NL(1)

    print("Hovedkatalog er:", hovedkatalog)
    NL(1)

    print("Excel-filnavn er:", Excel_filnavn)
    NL(1)

    path_til_Excel = os.path.normpath(os.path.join(hovedkatalog, Excel_filnavn))
    print("Excel-filbane er:", path_til_Excel)
    NL(1)

    print("Kontonummere er:", kontonummere)
    CSV_filnavn = [kontonummer + ".csv" for kontonummer in kontonummere]
    print("Og CSV-filnavn er:", CSV_filnavn)

    print(
        "Ekskluderer transaksjoner med disse beskrivelsene:",
        eksluder_disse_beskrivelser,
    )
    NL(1)

    oppdater_fra_dato = (
        datetime.now() - timedelta(days=antall_dager_tilbake)
    ).strftime("%Y-%m-%d")
    oppdater_til_dato = datetime.today().strftime("%Y-%m-%d")
    print(
        "Oppdaterer data fra:",
        oppdater_fra_dato,
        "til:",
        oppdater_til_dato,
        "altså",
        antall_dager_tilbake,
        "dager tilbake",
    )
    NL(1)

    if pauseperiode_i_timer <= 1:
        print(
            "Dato-kodet pauseperiode mellom hver kjøring er satt til:",
            pauseperiode_i_timer,
            "time",
        )
        NL(1)
    else:
        print(
            "Dato-kodet pauseperiode mellom hver kjøring er satt til:",
            pauseperiode_i_timer,
            "timer",
        )
        NL(1)

    print("Startsaldo før tidligste transaksjon i regnearket er satt til:", startsaldo)
    NL(1)

    print("REFRESH_TOKEN_FILE er:", REFRESH_TOKEN_FILE)
    NL(1)
    print("ACCESS_TOKEN_FILE er:", ACCESS_TOKEN_FILE)
    NL(1)
    print("AccountKeyMap_Fil er:", AccountKeyMap_Fil)
    NL(1)

    time.sleep(3)
    NL(5)
    print(
        "----------------------------------------------------------------------------------------------------------------------------------"
    )
    print(
        "---------------------------------------------------- Starter program nummer",
        telle_variabel,
        "----------------------------------------------------",
    )
    print(
        "----------------------------------------------------------------------------------------------------------------------------------"
    )
    NL(2)

    return


def lukk_arbeidsbok_global():
    arbeidsbok.close()
    # print("Arbeidsbok lukket ..."); NL(1)


def bare_1_gang():
    print("------------------------------------------------------")
    print("Programmet er nå kjørt 1 gang og avsluttes normalt ...")
    print("------------------------------------------------------")
    NL(3)
    exit()


def bare_3_ganger():
    # global telle_variabel
    # telle_variabel += 1
    if telle_variabel == 3:
        print("--------------------------------------------------------")
        print("Programmet er nå kjørt 3 ganger og avsluttes normalt ...")
        print("--------------------------------------------------------")
        NL(3)
        exit()


def NL(linjer):
    print("\n" * linjer)


# --------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# KONFIGURASJON


telle_variabel = 0

while True:


    # Konfigurering 1:
    #H_katalog = "/home/andreas/Downloads/WindowsBackup-main/Arkiv/Økonomi/H (Tom)/"
    H_katalog = "/media/andreas/UBUNTU 24_0/FlashdriveData/H/"
    # H_katalog = 'C:/Users/andre/H/'

    # Konfigurering 2:
    path_til_CSV = "/home/andreas/ws/config/budsjett/SB1_Data"
    # path_til_CSV = 'C:/Users/andre/OneDrive/Arkiv/Økonomi/SB1_Data/'

    # Konfigurering 3:
    hovedkatalog = "/home/andreas/Downloads/"
    # hovedkatalog = 'C:/Users/andre/OneDrive/Arkiv/Økonomi/'

    # Konfigurering 4:
    Excel_filnavn = "Budsjett.xlsx"
    arbeidsbok = last_arbeidsbok_global()
    NL(2)

    # Kondifuering 5:
    arbeidsark = arbeidsbok["Transaksjoner"]

    # Konfigurering 6:
    kontonummere = [
        "18131796605",
        "18133745544",
        "18135700327",
        "18227346139",
        "18138191555",
    ]

    # Konfigurering 7:
    eksluder_disse_beskrivelser = [
        "Jan Andreas Salvesen",
        "Overføring",
        "Nettbank til: 1813.37.45544",
    ]

    # Konfigurering 8:
    antall_dager_tilbake = 150

    # Konfigurering 9:
    antall_nye_kolonner = 15

    # Konfigurering 10:
    pauseperiode_i_timer = 8

    # Konfigurering 11:
    startsaldo = 39849

    # Konfigurering 12:
    REFRESH_TOKEN_FILE = os.path.join(H_katalog, "refresh_token.txt")
    # + Legg inn refresh_token i filen 'refresh_token.txt' i H-katalogen

    # Konfigurering 13:
    ACCESS_TOKEN_FILE = os.path.join(H_katalog, "access_token.txt")
    # + Legg inn access_token i filen 'access_token.txt' i H-katalogen

    # KOnfigurering 14:
    AccountKeyMap_Fil = os.path.join(H_katalog, "account_key-map")
    # + Legg inn account_key-map i filen 'account_key-map' i H-katalogen

    telle_variabel += 1
    printStartInfo()

    # --------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
    # Programstyring og valg av bank

    def Sparebank1():

        hent_transaksjonerSB1()
        time.sleep(4)

        les_og_sorter()
        time.sleep(4)

        # fjern_all_data_BtilG(); time.sleep(4)
        fjern_overlappende_data_BtilG()
        time.sleep(4)

        les_og_sorter()
        time.sleep(4)

        hent_kategoriser_og_skriv_CSV()
        time.sleep(4)

        les_og_sorter()
        time.sleep(4)

        skriv_predefinerte_transaksjoner()
        time.sleep(4)

        beregn_og_skriv_saldo()
        time.sleep(4)

        skriv_transaksjonstyper()
        time.sleep(4)

        skriv_endringstid()
        time.sleep(4)

        lagre()

        # bare_1_gang()
        # bare_3_ganger()

    def DNB():
        print("DNB er ikke implementert enda")
        NL(1)
        exit()

    def Nordea():
        print("Nordea er ikke implementert enda")
        NL(1)
        exit()

    def DanskeBank():
        print("Danske Bank er ikke implementert enda")
        NL(1)
        exit()

    def Handelsbanken():
        print("Handelsbanken er ikke implementert enda")
        NL(1)
        exit()

    def Sbanken():
        print("Sbanken er ikke implementert enda")
        NL(1)
        exit()

    Sparebank1()

    # DNB()
    # Nordea()
    # DanskeBank()
    # Handelsbanken()
    # Sbanken()

    regnskapforing()
